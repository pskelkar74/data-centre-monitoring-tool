from flask import Flask, render_template, request, json, url_for, send_file
import sys
import os
import happybase
import openpyxl as xl
from datetime import datetime
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
app.config['SECRET_KEY'] = 'secret'


def getPhysicalPosition(ip_addr):
    print(os.path.join(sys.path[0], 'input.xlsx'))
    wb = xl.load_workbook(os.path.join(sys.path[0], 'input.xlsx'))
    ws = wb.active
    cols = list(map(chr, range(65, 65 + ws.max_column)))
    rows = list(map(str, range(2, ws.max_row + 1)))
    inputdata = dict()

    for c in cols:
        inputdata[str(ws[c + "1"].value).strip()] = list()
    for c in cols:
        for r in rows:
            inputdata[str(ws[c + "1"].value).strip()
                      ].append(str(ws[c + r].value).strip())

    for i in range(len(inputdata["IP"])):
        if(inputdata["IP"][i] == ip_addr):
            rack_number = inputdata["Rack_number"][i]
            server_position = inputdata["System_number"][i]
            wb.close()
            return [rack_number, server_position]
    return []

def to_string(data):
    return ''.join(chr(x) for x in bytearray(data))


def to_list(data):
    return data.strip('()').split(',')


@app.route('/')
def index():
    return render_template('table.html')


@app.route('/api/v1/snmp', methods={'GET'})
def get_snmp():
    try:
        conn = happybase.Connection('localhost', port=9090)
        table = conn.table('snmp')
        row_key = str(datetime.now())[:16]
        conn.open()

        data = [data for (key, data) in table.scan(row_start=row_key)]
        conn.close()

        response = {}

        for x in data[-1]:
            key = to_string(x)
            value = to_string(data[-1][x])
            ip_addr = key.split(':')[0]
            if(ip_addr not in response.keys()):
                response[ip_addr] = {}
                position = getPhysicalPosition(ip_addr)
                response[ip_addr]["rack_pos"] = position[0]
                response[ip_addr]["server_pos"] = position[1]

            if(key.split(':')[1] == 'cpu'):
                tmp = to_list(value)
                # response[ip_addr].append(tmp[0])
                response[ip_addr]["cpu"] = tmp[0]

            elif(key.split(':')[1] == 'disk'):
                tmp = value.strip('[]').split('), ')
                physical_mem = to_list(tmp[0])
                virtual_mem = to_list(tmp[1])
                mem_buffers = to_list(tmp[2])
                # used physical space / total physical space *100
                perc = round((float(physical_mem[3])/float(physical_mem[2]))*100)
                response[ip_addr]["disk"] = perc

                # used virtual mem / total virtual mem *100
                perc = round((float(virtual_mem[3])/float(virtual_mem[2]))*100)
                response[ip_addr]["virtual_mem"] = perc

                # used memory buffer / total memory buffer *100
                perc = round((float(mem_buffers[3])/float(mem_buffers[2]))*100)
                response[ip_addr]["mem_buffers"] = perc

            elif(key.split(':')[1] == 'memory'):
                tmp = to_list(value)
                # used swap space / total swap space *100
                perc = round((float(tmp[1]) / float(tmp[0]))*100)
                response[ip_addr]["memory"] = perc
            elif(key.split(':')[1] == 'os'):
                os = value.split()[0]
                full_os = " ".join(value.split()[:3])
                response[ip_addr]["os"] = os
                response[ip_addr]["full_os"] = full_os
            elif(key.split(':')[1] == 'upt'):
                uptime = float(value)/3600
                response[ip_addr]["upt"] = str(round(uptime))
            # print(x, data[-1][x])
            # print('\n\n\n')

        '''
            Response format {'<ipAddr>': {'rack_pos': '1', 'server_pos': '1', 'cpu': '15', 'disk': 97, 'memory': 92, 'os': 'Linux', 'full_os': 'Linux rrk-lenovo 5.3.0-53-generic #47~18.04.1-Ubuntu SMP Thu May 7 13:10:50 UTC 2020 x86_64', 'upt': '506'}}
        '''
        return json.dumps(response), 200, {'ContentType': 'application/json'}
    except:
        return json.dumps({}), 400, {'ContentType': 'application/json'}


@app.route('/api/v1/ping', methods={'GET'})
def get_ping():
    try:
        conn = happybase.Connection('localhost', port=9090)
        table = conn.table('ping')
        row_key = str(datetime.now())[:16]
        conn.open()

        data = [data for (key, data) in table.scan(row_start=row_key)]
        conn.close()

        response = {}

        for x in data[-1]:
            key = to_string(x)
            value = to_string(data[-1][x])
            ip_addr = key.split(':')[0]
            if(ip_addr not in response.keys()):
                response[ip_addr] = ""

            if(key.split(':')[1] == 'ping'):
                response[ip_addr] = value

        '''
            Response format {<ip_addr>: "True" <or> "False", ...}
        '''
        return json.dumps(response), 200, {'ContentType': 'application/json'}
    except:
        return json.dumps({}), 400, {'ContentType': 'application/json'}


@app.route('/api/v1/ping/<ip_addr>', methods={'GET'})
def get_ip_ping(ip_addr):
    try:
        conn = happybase.Connection('localhost', port=9090)
        table = conn.table('ping')
        row_key = str(datetime.now())[:16]
        conn.open()

        data = [data for (key, data) in table.scan(row_start=row_key)]
        conn.close()

        response = {}

        for x in data[-1]:
            key = to_string(x)
            value = to_string(data[-1][x])
            ip = key.split(':')[0]
            if(ip == ip_addr):
                response["data"] = value
                break

        '''
            Response format { data: "True" <or> "False", ...}
        '''
        return json.dumps(response), 200, {'ContentType': 'application/json'}
    except:
        return json.dumps({}), 400, {'ContentType': 'application/json'}


@app.route('/api/v1/ssh', methods={'GET'})
def get_ssh():
    try:
        conn = happybase.Connection('localhost', port=9090)
        table = conn.table('ssh')
        row_key = str(datetime.now())[:16]
        conn.open()

        data = [data for (key, data) in table.scan(row_start=row_key)]
        conn.close()

        response = {}

        for x in data[-1]:
            key = to_string(x)
            value = to_string(data[-1][x])
            ip_addr = key.split(':')[0]
            if(ip_addr not in response.keys()):
                response[ip_addr] = {}

            if(key.split(':')[1] == 'ssh'):
                response[ip_addr]["sshStatus"] = value
            elif(key.split(':')[1] == 'last'):
                value = value.rstrip("\n")
                response[ip_addr]["last"] = value

        '''
            Response format {<ip_addr>: {"last": "user1\nuser1\nuser1\nuser1\nuser1", "sshStatus": "True" or "False"}}

        '''
        return json.dumps(response), 200, {'ContentType': 'application/json'}
    except:
        return json.dumps({}), 400, {'ContentType': 'application/json'}


@app.route('/api/v1/ssh/<ip_addr>', methods={'GET'})
def get_ip_ssh(ip_addr):
    try:
        conn = happybase.Connection('localhost', port=9090)
        table = conn.table('ssh')
        row_key = str(datetime.now())[:16]
        conn.open()

        data = [data for (key, data) in table.scan(row_start=row_key)]
        conn.close()

        response = {}

        for x in data[-1].keys():
            key = to_string(x)
            value = to_string(data[-1][x])
            ip = key.split(':')[0]

            if(ip == ip_addr):
                if(ip_addr not in response.keys()):
                    response[ip_addr] = {}

                if(key.split(':')[1] == 'ssh'):
                    response[ip_addr]["sshStatus"] = value
                elif(key.split(':')[1] == 'last'):
                    value = value.rstrip("\n")
                    response[ip_addr]["last"] = value

                if("sshStatus" in response.keys() and "last" in response.keys()):
                    break

        '''
            Response format {<ip_addr>: {"last": "user1\nuser1\nuser1\nuser1\nuser1", "sshStatus": "True" or "False"}}
        '''
        return json.dumps(response), 200, {'ContentType': 'application/json'}
    except:
        return json.dumps({}), 400, {'ContentType': 'application/json'}


@app.route('/download')
def download():
    path = 'diskdata.txt'
    return send_file(path)


if __name__ == '__main__':
    app.run(debug=True)
